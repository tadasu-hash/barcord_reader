from __future__ import annotations

import csv
import io
import json
import sqlite3
from openpyxl import load_workbook
from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import HTMLResponse
from pydantic import BaseModel, Field

DB_PATH = Path("asset_manager.db")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with closing(get_conn()) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS assets (
                asset_id TEXT PRIMARY KEY,
                barcode TEXT NOT NULL UNIQUE,
                asset_name TEXT NOT NULL,
                category TEXT,
                location TEXT,
                registered_at TEXT NOT NULL,
                last_confirmed_at TEXT,
                status TEXT NOT NULL DEFAULT 'active',
                is_deleted INTEGER NOT NULL DEFAULT 0,
                deleted_at TEXT,
                deleted_by TEXT,
                delete_comment TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS scan_logs (
                scan_log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                scanned_code TEXT NOT NULL,
                result TEXT NOT NULL,
                asset_id TEXT,
                message TEXT,
                scanned_at TEXT NOT NULL,
                scanned_by TEXT
            );

            CREATE TABLE IF NOT EXISTS audit_logs (
                audit_log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                target_asset_id TEXT,
                before_json TEXT,
                after_json TEXT,
                acted_at TEXT NOT NULL,
                acted_by TEXT,
                comment TEXT
            );
            """
        )
        conn.commit()


class AssetCreate(BaseModel):
    barcode: str = Field(min_length=1)
    asset_name: str = Field(min_length=1)
    registered_at: str
    category: str | None = None
    location: str | None = None
    status: str = "active"
    acted_by: str | None = None


class AssetUpdate(BaseModel):
    asset_name: str | None = None
    category: str | None = None
    location: str | None = None
    status: str | None = None
    acted_by: str | None = None


class ScanRequest(BaseModel):
    barcode: str
    scanned_by: str | None = None


class DeleteRequest(BaseModel):
    deleted_by: str | None = None
    delete_comment: str = Field(min_length=1)


app = FastAPI(title="資産バーコード管理アプリ")
init_db()


@app.get("/", response_class=HTMLResponse)
def japanese_gui() -> str:
    return """
<!doctype html><html lang="ja"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>資産バーコード管理</title>
<style>
body{font-family:sans-serif;max-width:920px;margin:20px auto;padding:0 12px;line-height:1.5}
.card{border:1px solid #ddd;border-radius:10px;padding:14px;margin-bottom:14px}
input,button{font-size:16px;padding:8px;margin:4px 0}
input{width:100%;max-width:420px}.row{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.ok{color:#0a7}.err{color:#c33} table{width:100%;border-collapse:collapse} th,td{border:1px solid #ddd;padding:6px;font-size:14px}
</style></head><body>
<h1>資産バーコード管理（かんたん画面）</h1>
<p>この画面は日本語で、一般ユーザー向けに最小操作だけに絞っています。</p>

<div class="card"><h2>1) バーコード読み取り（手入力でも可）</h2>
<input id="scanBarcode" placeholder="例: ABC-000123" autofocus onkeydown="onScanKey(event)"><button onclick="scan()">読み取る</button>
<div id="scanResult"></div></div>

<div class="card"><h2>2) 資産を登録</h2>
<div class="row"><input id="newBarcode" placeholder="バーコード"><input id="newName" placeholder="資産名"></div>
<div class="row"><input id="newDate" placeholder="登録日 例: 2026-05-19"><input id="newLocation" placeholder="設置場所（任意）"></div>
<button onclick="createAsset()">登録する</button><div id="createResult"></div></div>

<div class="card"><h2>3) 資産検索</h2>
<div class="row"><input id="qName" placeholder="資産名（部分一致）"><input id="qBarcode" placeholder="バーコード（完全一致）"></div>
<div class="row"><input id="qConfirmedFrom" type="date" placeholder="照合日From"><input id="qConfirmedTo" type="date" placeholder="照合日To"></div>
<label><input id="qUnconfirmed" type="checkbox"> 未確認資産のみ</label>
<label><input id="qNotConfirmedInPeriod" type="checkbox"> 指定期間に未照合のみ</label><br>
<button onclick="searchAssets()">検索する</button>
<div id="searchResult"></div></div>

<script>
function show(id, msg, ok=true){document.getElementById(id).innerHTML='<p class="'+(ok?'ok':'err')+'">'+msg+'</p>'}
function focusScanInput(){const el=document.getElementById('scanBarcode');el.focus();el.select();}
function onScanKey(event){if(event.key==='Enter'){event.preventDefault();scan();}}
async function scan(){
 const barcode=document.getElementById('scanBarcode').value.trim();
 if(!barcode){show('scanResult','バーコードを入力してください',false);return;}
 const r=await fetch('/api/v1/scan',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({barcode})});
 const d=await r.json();
 if(!r.ok){show('scanResult','未登録です: '+(d.detail?.message||'エラー'),false);focusScanInput();return;}
 show('scanResult','確認OK: '+d.asset.asset_name+'（最終確認: '+d.asset.last_confirmed_at+'）');
 document.getElementById('scanBarcode').value='';
 focusScanInput();
}

async function createAsset(){
 const payload={barcode:newBarcode.value.trim(),asset_name:newName.value.trim(),registered_at:newDate.value.trim(),location:newLocation.value.trim()||null};
 const r=await fetch('/api/v1/assets',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
 const d=await r.json();
 if(!r.ok){show('createResult','登録失敗: '+(d.detail||'エラー'),false);return;}
 show('createResult','登録しました（ID: '+d.asset_id+'）');
}
async function searchAssets(){
 const p=new URLSearchParams();
 if(qName.value.trim())p.set('asset_name',qName.value.trim());
 if(qBarcode.value.trim())p.set('barcode',qBarcode.value.trim());
 if(qUnconfirmed.checked)p.set('unconfirmed_only','true');
 if(qConfirmedFrom.value)p.set('confirmed_from',qConfirmedFrom.value);
 if(qConfirmedTo.value)p.set('confirmed_to',qConfirmedTo.value);
 if(qNotConfirmedInPeriod.checked)p.set('not_confirmed_in_period','true');
 const r=await fetch('/api/v1/assets?'+p.toString());
 const d=await r.json();
 if(!r.ok){show('searchResult','検索失敗',false);return;}
 let html='<p>'+d.count+'件</p><table><tr><th>バーコード</th><th>資産名</th><th>場所</th><th>最終確認</th></tr>';
 for(const x of d.items){html+=`<tr><td>${x.barcode||''}</td><td>${x.asset_name||''}</td><td>${x.location||''}</td><td>${x.last_confirmed_at||'未確認'}</td></tr>`}
 html+='</table>';document.getElementById('searchResult').innerHTML=html;
}
window.addEventListener('load',()=>focusScanInput());
</script></body></html>
"""


def audit(action: str, target_asset_id: str | None, before_json: str | None, after_json: str | None, acted_by: str | None, comment: str | None = None) -> None:
    with closing(get_conn()) as conn:
        conn.execute(
            """INSERT INTO audit_logs(action, target_asset_id, before_json, after_json, acted_at, acted_by, comment)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (action, target_asset_id, before_json, after_json, now_iso(), acted_by, comment),
        )
        conn.commit()


@app.post("/api/v1/scan")
def scan_asset(req: ScanRequest) -> dict[str, Any]:
    code = req.barcode.strip()
    with closing(get_conn()) as conn:
        row = conn.execute("SELECT * FROM assets WHERE barcode = ? AND is_deleted = 0", (code,)).fetchone()
        if row is None:
            conn.execute("INSERT INTO scan_logs(scanned_code, result, message, scanned_at, scanned_by) VALUES (?, 'not_found', ?, ?, ?)", (code, "asset not found for barcode", now_iso(), req.scanned_by))
            conn.commit()
            raise HTTPException(status_code=404, detail={"result": "not_found", "message": "asset not found for barcode"})
        ts = now_iso()
        conn.execute("UPDATE assets SET last_confirmed_at = ?, updated_at = ? WHERE asset_id = ?", (ts, ts, row["asset_id"]))
        conn.execute("INSERT INTO scan_logs(scanned_code, result, asset_id, message, scanned_at, scanned_by) VALUES (?, 'success', ?, ?, ?, ?)", (code, row["asset_id"], "matched", ts, req.scanned_by))
        conn.commit()
    return {"result": "success", "asset": {"asset_id": row["asset_id"], "barcode": row["barcode"], "asset_name": row["asset_name"], "location": row["location"], "last_confirmed_at": ts}}


@app.get("/api/v1/assets")
def list_assets(
    asset_name: str | None = None,
    barcode: str | None = None,
    unconfirmed_only: bool = False,
    confirmed_from: str | None = None,
    confirmed_to: str | None = None,
    not_confirmed_in_period: bool = False,
) -> dict[str, Any]:
    query = "SELECT * FROM assets WHERE is_deleted = 0"
    params: list[Any] = []
    if asset_name:
        query += " AND asset_name LIKE ?"
        params.append(f"%{asset_name}%")
    if barcode:
        query += " AND barcode = ?"
        params.append(barcode)
    if unconfirmed_only:
        query += " AND last_confirmed_at IS NULL"
    if confirmed_from and confirmed_to and not_confirmed_in_period:
        query += " AND (last_confirmed_at IS NULL OR date(last_confirmed_at) < date(?) OR date(last_confirmed_at) > date(?))"
        params.extend([confirmed_from, confirmed_to])
    else:
        if confirmed_from:
            query += " AND date(last_confirmed_at) >= date(?)"
            params.append(confirmed_from)
        if confirmed_to:
            query += " AND date(last_confirmed_at) <= date(?)"
            params.append(confirmed_to)
    with closing(get_conn()) as conn:
        rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    return {"count": len(rows), "items": rows}


@app.post("/api/v1/assets")
def create_asset(req: AssetCreate) -> dict[str, Any]:
    asset_id = str(uuid4())
    ts = now_iso()
    with closing(get_conn()) as conn:
        try:
            conn.execute("""INSERT INTO assets(asset_id, barcode, asset_name, category, location, registered_at, status, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""", (asset_id, req.barcode.strip(), req.asset_name.strip(), req.category, req.location, req.registered_at, req.status, ts, ts))
            conn.commit()
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="barcode already exists")
    audit("create", asset_id, None, json.dumps(req.model_dump(), ensure_ascii=False), req.acted_by)
    return {"asset_id": asset_id}


@app.put("/api/v1/assets/{asset_id}")
def update_asset(asset_id: str, req: AssetUpdate) -> dict[str, str]:
    with closing(get_conn()) as conn:
        before = conn.execute("SELECT * FROM assets WHERE asset_id = ? AND is_deleted = 0", (asset_id,)).fetchone()
        if before is None:
            raise HTTPException(status_code=404, detail="asset not found")
        updates = {k: v for k, v in req.model_dump().items() if v is not None and k != "acted_by"}
        if not updates:
            return {"result": "no_change"}
        updates["updated_at"] = now_iso()
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        conn.execute(f"UPDATE assets SET {set_clause} WHERE asset_id = ?", list(updates.values()) + [asset_id])
        conn.commit()
    audit("update", asset_id, json.dumps(dict(before), ensure_ascii=False), json.dumps(updates, ensure_ascii=False), req.acted_by)
    return {"result": "updated"}


@app.delete("/api/v1/assets/{asset_id}")
def delete_asset(asset_id: str, req: DeleteRequest) -> dict[str, str]:
    ts = now_iso()
    with closing(get_conn()) as conn:
        row = conn.execute("SELECT * FROM assets WHERE asset_id = ? AND is_deleted = 0", (asset_id,)).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="asset not found")
        conn.execute("""UPDATE assets SET is_deleted = 1, deleted_at = ?, deleted_by = ?, delete_comment = ?, updated_at = ? WHERE asset_id = ?""", (ts, req.deleted_by, req.delete_comment, ts, asset_id))
        conn.commit()
    audit("delete", asset_id, json.dumps(dict(row), ensure_ascii=False), None, req.deleted_by, req.delete_comment)
    return {"result": "deleted"}


@app.post("/api/v1/assets/import")
def import_assets(file: UploadFile = File(...)) -> dict[str, Any]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="missing filename")

    inserted = 0
    errors: list[dict[str, Any]] = []

    if file.filename.lower().endswith(".csv"):
        content = file.file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        required = {"barcode", "asset_name", "registered_at"}
        if not required.issubset(set(reader.fieldnames or [])):
            raise HTTPException(status_code=422, detail=f"required columns: {sorted(required)}")

        for idx, row in enumerate(reader, start=2):
            barcode = (row.get("barcode") or "").strip()
            asset_name = (row.get("asset_name") or "").strip()
            registered_at = (row.get("registered_at") or "").strip()
            if not barcode or not asset_name or not registered_at:
                errors.append({"row": idx, "error": "missing required values"})
                continue
            try:
                create_asset(AssetCreate(barcode=barcode, asset_name=asset_name, registered_at=registered_at, category=row.get("category"), location=row.get("location")))
                inserted += 1
            except HTTPException as exc:
                errors.append({"row": idx, "error": str(exc.detail)})
        return {"inserted": inserted, "errors": errors}

    if file.filename.lower().endswith(".xlsx"):
        wb = load_workbook(filename=io.BytesIO(file.file.read()), read_only=True, data_only=True)
        ws = wb.active
        # F列=固定資産番号(バーコード), G列=枝番, H列=資産名称, J列=事業所名
        # 要望に合わせて J列が「食品研究所」の行だけを取り込み。
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=10, values_only=True), start=2):
            barcode = str(row[5]).strip() if row[5] is not None else ""
            asset_name = str(row[6]).strip() if row[6] is not None else ""
            office_name = str(row[9]).strip() if row[9] is not None else ""
            if office_name != "食品研究所":
                continue
            if not barcode or not asset_name:
                errors.append({"row": row_idx, "error": "missing required values at F or G"})
                continue
            try:
                create_asset(
                    AssetCreate(
                        barcode=barcode,
                        asset_name=asset_name,
                        registered_at=now_iso()[:10],
                        location=office_name,
                    )
                )
                inserted += 1
            except HTTPException as exc:
                errors.append({"row": row_idx, "error": str(exc.detail)})
        return {"inserted": inserted, "errors": errors}

    raise HTTPException(status_code=422, detail="supported formats: .csv, .xlsx")


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}
